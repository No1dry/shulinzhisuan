import os
import json
import uuid
from datetime import datetime
from typing import Optional
from fastapi import APIRouter, Depends, Query
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.database import get_db
from app.models import ResidentMaster, Housing, ProblemReport
from app.schemas import ResponseModel
from app.services.header_recognizer import LLMReportGenerator
from app.encryption import decrypt_field

router = APIRouter(prefix="/reports", tags=["Reports"])

REPORTS_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "reports")
os.makedirs(REPORTS_DIR, exist_ok=True)

@router.post("/generate", response_model=ResponseModel)
async def generate_report(payload: dict, db: Session = Depends(get_db)):
    """Generate report"""
    try:
        report_type = payload.get("report_type", "population_summary")
        grid_name = payload.get("grid_name", None)
        filters = payload.get("filters", {})
        format_type = payload.get("format", "excel")
        
        # Gather statistics
        stats = {}
        
        base_query = db.query(ResidentMaster)
        if grid_name:
            base_query = base_query.filter(ResidentMaster.grid_name == grid_name)
        
        stats["total_residents"] = base_query.count()
        stats["key_populations"] = base_query.filter(ResidentMaster.is_key_population == True).count()
        stats["elderly_alone"] = base_query.filter(
            ResidentMaster.is_living_alone == True,
            ResidentMaster.age >= 60
        ).count()
        stats["disabled_persons"] = base_query.filter(ResidentMaster.is_disabled == True).count()
        stats["low_income"] = base_query.filter(ResidentMaster.is_low_income == True).count()
        stats["left_behind_children"] = base_query.filter(ResidentMaster.is_left_behind_child == True).count()
        stats["elderly_60"] = base_query.filter(ResidentMaster.age >= 60).count()
        stats["elderly_80"] = base_query.filter(ResidentMaster.age >= 80).count()
        
        grids = db.query(ResidentMaster.grid_name).distinct().filter(
            ResidentMaster.grid_name.isnot(None)
        ).count()
        stats["total_grids"] = grids
        
        # Grid breakdown
        grid_stats = []
        if report_type == "grid_statistics":
            grid_names = db.query(ResidentMaster.grid_name).distinct().filter(
                ResidentMaster.grid_name.isnot(None)
            ).all()
            for g in grid_names:
                if not g[0]:
                    continue
                count = db.query(ResidentMaster).filter(ResidentMaster.grid_name == g[0]).count()
                key_count = db.query(ResidentMaster).filter(
                    ResidentMaster.grid_name == g[0],
                    ResidentMaster.is_key_population == True
                ).count()
                elderly_count = db.query(ResidentMaster).filter(
                    ResidentMaster.grid_name == g[0],
                    ResidentMaster.age >= 60
                ).count()
                grid_stats.append({
                    "grid_name": g[0],
                    "resident_count": count,
                    "key_population_count": key_count,
                    "elderly_count": elderly_count
                })
        
        # Generate summary with LLM Mock
        summary = LLMReportGenerator.generate_summary(report_type, stats)
        
        # Generate report file
        report_id = str(uuid.uuid4())
        filename = f"report_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        if format_type == "excel":
            filename += ".xlsx"
            file_path = os.path.join(REPORTS_DIR, filename)
            
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                wb = openpyxl.Workbook()
                
                # Summary sheet
                ws_summary = wb.active
                ws_summary.title = "汇总数据"
                
                # Header style
                header_fill = PatternFill(start_color="2B5C8F", end_color="2B5C8F", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True, size=12)
                
                # Title
                ws_summary.merge_cells("A1:F1")
                ws_summary["A1"] = f"社区治理数据报告 - {report_type}"
                ws_summary["A1"].font = Font(size=16, bold=True, color="2B5C8F")
                ws_summary["A1"].alignment = Alignment(horizontal="center")
                
                # Summary text
                ws_summary.merge_cells("A2:F2")
                ws_summary["A2"] = summary
                ws_summary["A2"].alignment = Alignment(wrap_text=True)
                
                # Stats headers
                headers = ["指标", "数值"]
                for i, h in enumerate(headers):
                    cell = ws_summary.cell(row=4, column=i+1, value=h)
                    cell.fill = header_fill
                    cell.font = header_font
                
                # Stats data
                stat_items = [
                    ("居民总数", stats["total_residents"]),
                    ("网格数", stats["total_grids"]),
                    ("重点人群", stats["key_populations"]),
                    ("独居老人", stats["elderly_alone"]),
                    ("残疾人", stats["disabled_persons"]),
                    ("低保户", stats["low_income"]),
                    ("留守儿童", stats["left_behind_children"]),
                    ("60岁以上老人", stats["elderly_60"]),
                    ("80岁以上老人", stats["elderly_80"]),
                ]
                
                for i, (label, value) in enumerate(stat_items):
                    ws_summary.cell(row=5+i, column=1, value=label)
                    ws_summary.cell(row=5+i, column=2, value=value)
                
                # Grid breakdown sheet
                if grid_stats:
                    ws_grids = wb.create_sheet("网格分布")
                    grid_headers = ["网格名称", "居民人数", "重点人群数", "老人数"]
                    for i, h in enumerate(grid_headers):
                        cell = ws_grids.cell(row=1, column=i+1, value=h)
                        cell.fill = header_fill
                        cell.font = header_font
                    
                    for i, g in enumerate(grid_stats):
                        ws_grids.cell(row=2+i, column=1, value=g["grid_name"])
                        ws_grids.cell(row=2+i, column=2, value=g["resident_count"])
                        ws_grids.cell(row=2+i, column=3, value=g["key_population_count"])
                        ws_grids.cell(row=2+i, column=4, value=g["elderly_count"])
                
                wb.save(file_path)
                
            except ImportError:
                # Fallback: create CSV
                filename = filename.replace(".xlsx", ".csv")
                file_path = os.path.join(REPORTS_DIR, filename)
                with open(file_path, "w", encoding="utf-8-sig") as f:
                    f.write("指标,数值\n")
                    for label, value in [
                        ("居民总数", stats["total_residents"]),
                        ("重点人群", stats["key_populations"]),
                        ("独居老人", stats["elderly_alone"]),
                        ("残疾人", stats["disabled_persons"]),
                        ("低保户", stats["low_income"]),
                    ]:
                        f.write(f"{label},{value}\n")
        
        else:
            # JSON format
            filename += ".json"
            file_path = os.path.join(REPORTS_DIR, filename)
            with open(file_path, "w", encoding="utf-8") as f:
                json.dump({
                    "report_type": report_type,
                    "generated_at": datetime.now().isoformat(),
                    "summary": summary,
                    "statistics": stats,
                    "grid_breakdown": grid_stats
                }, f, ensure_ascii=False, indent=2)
        
        return ResponseModel(
            code=200,
            message="报告生成成功",
            data={
                "report_id": report_id,
                "report_type": report_type,
                "filename": filename,
                "download_url": f"/api/reports/download/{filename}",
                "summary": summary,
                "statistics": stats,
                "grid_breakdown": grid_stats,
                "generated_at": datetime.now().isoformat()
            }
        )
    
    except Exception as e:
        return ResponseModel(code=500, message=f"报告生成失败: {str(e)}", data=None)


@router.get("/download/{filename}")
async def download_report(filename: str):
    """Download report file"""
    from fastapi.responses import FileResponse
    
    file_path = os.path.join(REPORTS_DIR, filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="报告文件不存在")
    
    return FileResponse(
        path=file_path,
        filename=filename,
        media_type="application/octet-stream"
    )


@router.get("/types", response_model=ResponseModel)
async def get_report_types():
    """Get available report types"""
    types = [
        {"value": "population_summary", "label": "人口汇总报告"},
        {"value": "grid_statistics", "label": "网格统计报告"},
        {"value": "key_populations", "label": "重点人群报告"},
        {"value": "elderly_care", "label": "养老关怀报告"},
        {"value": "disability_support", "label": "残疾人帮扶报告"},
        {"value": "low_income_assistance", "label": "低保救助报告"},
    ]
    return ResponseModel(code=200, message="success", data=types)
