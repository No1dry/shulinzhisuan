import os
import re
import csv
from typing import List, Dict, Any, Tuple, Optional


def detect_header_row(raw_rows: List[List[Any]]) -> int:
    """Smart header row detection. Returns 0-based index of header row."""
    if not raw_rows or len(raw_rows) < 1:
        return 0
    
    def row_score(row_idx: int) -> float:
        if row_idx >= len(raw_rows):
            return 0
        row = raw_rows[row_idx]
        
        filled_cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
        total_cells = len(row)
        fill_rate = len(filled_cells) / max(total_cells, 1)
        
        header_keywords = [
            '姓名', '性别', '身份证', '电话', '手机', '地址', '网格', '楼栋',
            '单元', '户号', '民族', '婚姻', '就业', '医保', '低保', '残疾',
            '独居', '留守', '重点', '街道', '社区', '小区', '房号', '门牌',
            '出生', '年龄', '户口', '户籍', 'name', 'gender', 'phone', 'address',
            'id', 'card', 'age', 'birth', '人口', '人员', '户主',
            '序号', '编号', '备注', '状态', '类型', '面积', '流动',
            '租赁', '矛盾', '纠纷', '诉求', '房主', '房东', '业主',
        ]
        
        title_keywords = ['表', '统计', '汇总', '报告', '台账', '清单', '名册', '目录']
        
        keyword_matches = sum(1 for cell in filled_cells 
                             for kw in header_keywords if kw.lower() in cell.lower())
        title_matches = sum(1 for cell in filled_cells 
                           for tw in title_keywords if tw in cell)
        
        avg_len = sum(len(c) for c in filled_cells) / max(len(filled_cells), 1) if filled_cells else 0
        
        score = 0
        score += fill_rate * 30
        score += min(keyword_matches * 10, 50)
        score -= title_matches * 15
        
        if avg_len > 15:
            score -= 20
        elif 2 <= avg_len <= 10:
            score += 10
        
        if len(filled_cells) <= 2 and len(raw_rows) > 1:
            score -= 30
        
        return max(score, 0)
    
    score0 = row_score(0)
    score1 = row_score(1) if len(raw_rows) > 1 else 0
    
    if score1 > score0 + 5:
        return 1
    
    return 0


def extract_community_from_filename(filename: str) -> Optional[str]:
    """Extract community name from filename."""
    name = os.path.splitext(filename)[0]
    patterns = [
        r'(.+街道.+社区)',
        r'(.+镇.+村)',
        r'(.+?(?:社区|村|街道))(?:人口|居民|人员|房屋|住房|人房|信息|基础|统计|表|台账|名册|名单)',
        r'(.+?)(?:人口|居民|人员|房屋|住房|人房|信息|基础|统计|表|台账|名册|名单)',
    ]
    
    for pat in patterns:
        match = re.search(pat, name)
        if match:
            result = match.group(1).strip()
            if len(result) >= 4:
                return result
    
    return None


class SheetData:
    """Represents data from a single Excel sheet."""
    def __init__(self, name: str, headers: List[str], rows: List[Dict[str, Any]], title: Optional[str] = None):
        self.name = name
        self.headers = headers
        self.rows = rows
        self.title = title
        self.row_count = len(rows)


def parse_all_sheets(file_path: str, file_type: str) -> List[SheetData]:
    """
    Parse ALL sheets from Excel file.
    For CSV, returns single sheet.
    """
    ext = file_type.lower()
    
    if ext == 'xlsx':
        return parse_xlsx_all_sheets(file_path)
    elif ext == 'xls':
        return parse_xls_all_sheets(file_path)
    elif ext == 'csv':
        headers, rows = parse_csv(file_path)
        return [SheetData(name='Sheet1', headers=headers, rows=rows)]
    else:
        raise ValueError(f"不支持的文件格式: {ext}")


def _smart_number_to_str(cell_value: Any) -> str:
    """
    Convert Excel cell value to string, handling numeric cells properly.
    Key fix: phone numbers and ID cards stored as numbers get .0 suffix from float.
    Example: 13412345678.0 -> "13412345678", not "13412345678.0"
    """
    if cell_value is None:
        return ""
    
    if isinstance(cell_value, float):
        # Check if it's effectively an integer (phone, ID card, age, etc.)
        if cell_value == int(cell_value):
            return str(int(cell_value))
        else:
            # It's a real float, keep as string but strip trailing .0
            s = str(cell_value)
            if '.' in s:
                s = s.rstrip('0').rstrip('.')
            return s
    
    if isinstance(cell_value, int):
        return str(cell_value)
    
    return str(cell_value).strip()


def _extract_sheet_data(raw_rows: List[List[Any]], sheet_name: str) -> Optional[SheetData]:
    """Extract data from raw rows of a single sheet."""
    if not raw_rows:
        return None
    
    header_idx = detect_header_row(raw_rows)
    
    title = None
    if header_idx > 0:
        title_cells = [str(c).strip() for c in raw_rows[0] if c is not None and str(c).strip()]
        if title_cells:
            title = ' | '.join(title_cells)
    
    header_row = raw_rows[header_idx]
    headers = []
    for i, cell in enumerate(header_row):
        if cell is not None and str(cell).strip():
            headers.append(str(cell).strip())
        else:
            headers.append(f"Column_{i}")
    
    rows: List[Dict[str, Any]] = []
    for row_data in raw_rows[header_idx + 1:]:
        row_dict: Dict[str, Any] = {}
        has_content = False
        for i, cell in enumerate(row_data):
            key = headers[i] if i < len(headers) else f"Column_{i}"
            # Use smart conversion to avoid .0 on phone numbers/IDs
            val = _smart_number_to_str(cell)
            row_dict[key] = val
            if val:
                has_content = True
        
        if has_content:
            rows.append(row_dict)
    
    if not rows:
        return None
    
    return SheetData(name=sheet_name, headers=headers, rows=rows, title=title)


def parse_xlsx_all_sheets(file_path: str) -> List[SheetData]:
    """Parse all sheets from xlsx file."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        result: List[SheetData] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            raw_rows: List[List[Any]] = []
            for row in ws.iter_rows():
                raw_row = [cell.value for cell in row]
                if any(v is not None for v in raw_row):
                    raw_rows.append(raw_row)
            
            sheet_data = _extract_sheet_data(raw_rows, sheet_name)
            if sheet_data and sheet_data.row_count > 0:
                result.append(sheet_data)
        
        wb.close()
        return result
    
    except ImportError:
        raise ImportError("缺少 openpyxl 库，请执行: pip install openpyxl")
    except Exception as e:
        error_msg = str(e)
        if "zip" in error_msg.lower():
            raise ValueError(
                "该文件看起来是 .xls 老格式但扩展名为 .xlsx。"
                "请尝试：1) 用Excel另存为.xlsx格式后重新上传；2) 或将扩展名改为.xls"
            )
        raise ValueError(f"解析 xlsx 文件失败: {error_msg}")


def parse_xls_all_sheets(file_path: str) -> List[SheetData]:
    """Parse all sheets from xls file."""
    try:
        import xlrd
        wb = xlrd.open_workbook(file_path)
        
        result: List[SheetData] = []
        for sheet_idx in range(wb.nsheets):
            ws = wb.sheet_by_index(sheet_idx)
            sheet_name = ws.name
            
            raw_rows: List[List[Any]] = []
            for row_idx in range(ws.nrows):
                raw_row = [ws.cell_value(row_idx, col_idx) for col_idx in range(ws.ncols)]
                if any(v is not None and str(v).strip() for v in raw_row):
                    raw_rows.append(raw_row)
            
            sheet_data = _extract_sheet_data(raw_rows, sheet_name)
            if sheet_data and sheet_data.row_count > 0:
                result.append(sheet_data)
        
        return result
    
    except ImportError:
        raise ImportError("缺少 xlrd 库，请执行: pip install xlrd")
    except Exception as e:
        raise ValueError(f"解析 xls 文件失败: {str(e)}")


def parse_csv(file_path: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    """Parse CSV file with encoding detection."""
    encodings = ['utf-8', 'gbk', 'gb2312', 'gb18030', 'latin-1']
    
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                sample = f.read(4096)
                f.seek(0)
                
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=',，\t')
                except:
                    dialect = csv.excel
                
                reader = csv.DictReader(f, dialect=dialect)
                headers = reader.fieldnames or []
                rows = []
                for row in reader:
                    cleaned = {k.strip() if k else f"Column_{i}": (v.strip() if v else "")
                              for i, (k, v) in enumerate(row.items())}
                    if any(v for v in cleaned.values()):
                        rows.append(cleaned)
                return headers, rows
        except UnicodeDecodeError:
            continue
        except Exception:
            continue
    
    raise ValueError("无法识别CSV文件编码，支持的编码: utf-8, gbk, gb2312, gb18030")


# Legacy single-sheet function for backward compatibility
def parse_file(file_path: str, file_type: str) -> Tuple[List[str], List[Dict[str, Any]], Optional[str]]:
    """Parse first sheet only (legacy)."""
    sheets = parse_all_sheets(file_path, file_type)
    if not sheets:
        return [], [], None
    first = sheets[0]
    return first.headers, first.rows, first.title


STANDARD_FIELD_MAP = {
    "name": ["姓名", "名字", "居民姓名", "人员姓名", "户主姓名", "业主姓名", "name"],
    "gender": ["性别", "男女", "gender", "sex"],
    "id_card": ["身份证号", "身份证号码", "身份证", "身分证号", "id card", "idcard", "证件号码", "居民身份证号", "公民身份号码"],
    "phone": ["手机号", "手机号码", "联系电话", "电话", "手机", "phone", "mobile", "联系方式", "电话号码", "住户电话"],
    "birth_date": ["出生日期", "生日", "出生年月", "birth date", "birthday", "dob"],
    "age": ["年龄", "岁数", "age", "现年", "周岁"],
    "residence_address": ["居住地址", "现住址", "住址", "地址", "实际居住地", "现居住地址", "住所", "address", "详细地址"],
    "household_address": ["户籍地址", "户口地址", "户籍地", "户口所在地", "户籍所在地"],
    "grid_name": ["所属网格", "网格", "责任网格", "管理网格", "grid", "网格名称", "社区网格", "辖区", "社区"],
    "building_unit": ["楼栋单元", "楼号", "楼栋", "单元", "楼座", "幢号", "房号"],
    "household_number": ["户号", "门牌号", "户编号"],
    "ethnicity": ["民族", "少数民族", "ethnicity", "nation"],
    "marital_status": ["婚姻状况", "婚姻", "婚否", "marital status"],
    "employment_status": ["就业情况", "工作状态", "职业", "employment", "job"],
    "medical_insurance": ["医保状态", "医疗保险", "医保", "参保状态"],
    "is_low_income": ["低保", "低保户", "最低生活保障", "低保状态", "是否低保"],
    "is_disabled": ["残疾", "残疾人", "残障", "残疾状况", "是否残疾"],
    "disability_type": ["残疾类别", "残疾类型", "残疾等级"],
    "is_living_alone": ["独居", "独居老人", "独自居住", "是否独居"],
    "is_left_behind_child": ["留守儿童", "留守", "是否留守"],
    "is_key_population": ["重点人群", "重点人员", "重点关注", "是否重点"],
    "key_population_type": ["重点人群类别", "重点人员类型", "重点类别"],
    "is_special_support": ["特困", "特困人员", "特困供养", "是否特困"],
    "relation_to_household": ["与户主关系"],
    "is_floating_population": ["是否流动人口", "流动人口"],
    "is_rental": ["是否租赁房屋", "租赁房屋"],
    "conflict_dispute": ["矛盾纠纷"],
    "opinion_appeal": ["意见诉求"],
    "street": ["街道"],
    "neighborhood": ["社区（村）", "社区"],
    "community_group": ["小区（组）", "小区"],
    "is_owner": ["是否房主", "房主", "房东"],
}

def standardize_field_name(header: str) -> str:
    """Try to match a header to a standard field name."""
    header_clean = header.strip().lower()
    
    for standard, variations in STANDARD_FIELD_MAP.items():
        for var in variations:
            if header_clean == var.lower():
                return standard
    
    for standard, variations in STANDARD_FIELD_MAP.items():
        for var in variations:
            if var.lower() in header_clean or header_clean in var.lower():
                return standard
    
    return ""
