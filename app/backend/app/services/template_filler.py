"""
Template table auto-fill service.
Upload a template (headers only), system fills resident static data.
"""
from typing import List, Dict, Any, Optional, Tuple
from sqlalchemy.orm import Session
from app.models import ResidentMaster
from app.encryption import decrypt_field
from app.services.file_parser import STANDARD_FIELD_MAP

# Reverse lookup: standard_field -> list of possible original headers
REVERSE_FIELD_MAP: Dict[str, List[str]] = {
    "name": ["姓名", "名字", "人员姓名", "居民姓名", "户主姓名", "业主姓名", "Name"],
    "gender": ["性别", "男女", "Gender", "Sex"],
    "id_card": ["身份证号", "身份证号码", "身份证", "证件号码", "ID Card", "id_card"],
    "phone": ["手机号", "手机号码", "联系电话", "电话", "手机", "Phone", "Mobile", "联系方式"],
    "birth_date": ["出生日期", "生日", "出生年月", "Birth Date"],
    "age": ["年龄", "岁数", "Age"],
    "ethnicity": ["民族", "Ethnicity"],
    "marital_status": ["婚姻状况", "婚姻", "婚否", "Marital Status"],
    "employment_status": ["就业情况", "工作状态", "职业", "Employment"],
    "medical_insurance": ["医保状态", "医疗保险", "医保", "参保状态"],
    "residence_address": ["居住地址", "现住址", "住址", "地址", "实际居住地", "详细地址", "Residence Address"],
    "household_address": ["户籍地址", "户口地址", "户籍地", "户籍所在地"],
    "grid_name": ["所属网格", "网格", "社区", "小区", "Grid", "Community"],
    "building_unit": ["楼栋单元", "楼号", "楼栋", "单元", "Building Unit"],
    "household_number": ["户号", "门牌号"],
    "is_low_income": ["是否低保", "低保", "低保户"],
    "is_disabled": ["是否残疾", "残疾", "残疾人"],
    "disability_type": ["残疾类别", "残疾类型"],
    "is_living_alone": ["是否独居", "独居"],
    "is_left_behind_child": ["是否留守儿童", "留守儿童"],
    "is_key_population": ["是否重点人群", "重点人群"],
    "key_population_type": ["重点人群类别"],
    "is_special_support": ["是否特困", "特困"],
    "relation_to_household": ["与户主关系"],
    "street": ["街道"],
    "neighborhood": ["社区(村)", "社区（村）"],
    "community_group": ["小区(组)", "小区（组）"],
}

# Model field -> attribute name on ResidentMaster
FIELD_TO_MODEL_ATTR: Dict[str, str] = {
    "name": "name_encrypted",
    "gender": "gender",
    "id_card": "id_card_encrypted",
    "phone": "phone_encrypted",
    "birth_date": "birth_date",
    "age": "age",
    "ethnicity": "ethnicity",
    "marital_status": "marital_status",
    "employment_status": "employment_status",
    "medical_insurance": "medical_insurance",
    "residence_address": "residence_address",
    "household_address": "household_address",
    "grid_name": "grid_name",
    "building_unit": "building_unit",
    "household_number": "household_number",
    "is_low_income": "is_low_income",
    "is_disabled": "is_disabled",
    "disability_type": "disability_type",
    "is_living_alone": "is_living_alone",
    "is_left_behind_child": "is_left_behind_child",
    "is_key_population": "is_key_population",
    "key_population_type": "key_population_type",
    "is_special_support": "is_special_support",
    "relation_to_household": None,  # from custom_fields
    "street": None,
    "neighborhood": None,
    "community_group": None,
}


def match_template_headers(template_headers: List[str]) -> List[Dict[str, Any]]:
    """
    Match template headers to standard resident fields.
    Returns list of {template_header, standard_field, display_name, matched}
    """
    results = []
    for header in template_headers:
        header_clean = header.strip().lower()
        matched_field = ""
        display_name = header
        
        # Try exact match first
        for std_field, aliases in REVERSE_FIELD_MAP.items():
            for alias in aliases:
                if header_clean == alias.lower():
                    matched_field = std_field
                    display_name = STANDARD_FIELD_MAP.get(std_field, [std_field])[0] if std_field in STANDARD_FIELD_MAP else std_field
                    break
            if matched_field:
                break
        
        # Fuzzy match
        if not matched_field:
            for std_field, aliases in REVERSE_FIELD_MAP.items():
                for alias in aliases:
                    if alias.lower() in header_clean or header_clean in alias.lower():
                        matched_field = std_field
                        display_name = aliases[0]
                        break
                if matched_field:
                    break
        
        results.append({
            "template_header": header,
            "standard_field": matched_field,
            "display_name": display_name,
            "matched": bool(matched_field)
        })
    
    return results


def get_resident_field_value(resident: ResidentMaster, standard_field: str) -> str:
    """Get a field value from a resident record, decrypting if necessary."""
    if standard_field == "name":
        return decrypt_field(resident.name_encrypted) if resident.name_encrypted else ""
    elif standard_field == "id_card":
        return decrypt_field(resident.id_card_encrypted) if resident.id_card_encrypted else ""
    elif standard_field == "phone":
        return decrypt_field(resident.phone_encrypted) if resident.phone_encrypted else ""
    elif standard_field == "is_low_income":
        return "是" if resident.is_low_income else "否"
    elif standard_field == "is_disabled":
        return "是" if resident.is_disabled else "否"
    elif standard_field == "is_living_alone":
        return "是" if resident.is_living_alone else "否"
    elif standard_field == "is_left_behind_child":
        return "是" if resident.is_left_behind_child else "否"
    elif standard_field == "is_key_population":
        return "是" if resident.is_key_population else "否"
    elif standard_field == "is_special_support":
        return "是" if resident.is_special_support else "否"
    elif standard_field == "age":
        return str(resident.age) if resident.age is not None else ""
    elif standard_field == "gender":
        return resident.gender or ""
    elif standard_field == "birth_date":
        return resident.birth_date or ""
    elif standard_field == "ethnicity":
        return resident.ethnicity or ""
    elif standard_field == "marital_status":
        return resident.marital_status or ""
    elif standard_field == "employment_status":
        return resident.employment_status or ""
    elif standard_field == "medical_insurance":
        return resident.medical_insurance or ""
    elif standard_field == "residence_address":
        return resident.residence_address or ""
    elif standard_field == "household_address":
        return resident.household_address or ""
    elif standard_field == "grid_name":
        return resident.grid_name or ""
    elif standard_field == "building_unit":
        return resident.building_unit or ""
    elif standard_field == "household_number":
        return resident.household_number or ""
    elif standard_field == "disability_type":
        return resident.disability_type or ""
    elif standard_field == "key_population_type":
        return resident.key_population_type or ""
    else:
        # Try custom_fields
        if resident.custom_fields and standard_field in resident.custom_fields:
            return str(resident.custom_fields[standard_field])
        return ""


def fill_template(
    db: Session,
    template_headers: List[str],
    field_mapping: List[Dict[str, str]],
    grid_name: Optional[str] = None,
    filters: Optional[Dict[str, Any]] = None
) -> List[Dict[str, str]]:
    """
    Fill template with resident data.
    field_mapping: list of {template_header, standard_field}
    Returns list of rows (dicts with template_header as keys).
    """
    # Query residents
    query = db.query(ResidentMaster)
    
    if grid_name:
        query = query.filter(ResidentMaster.grid_name == grid_name)
    
    if filters:
        if filters.get("is_key_population"):
            query = query.filter(ResidentMaster.is_key_population == True)
        if filters.get("is_living_alone"):
            query = query.filter(ResidentMaster.is_living_alone == True)
        if filters.get("is_disabled"):
            query = query.filter(ResidentMaster.is_disabled == True)
        if filters.get("is_low_income"):
            query = query.filter(ResidentMaster.is_low_income == True)
        if filters.get("age_min"):
            query = query.filter(ResidentMaster.age >= int(filters["age_min"]))
        if filters.get("age_max"):
            query = query.filter(ResidentMaster.age <= int(filters["age_max"]))
    
    residents = query.all()
    
    # Build mapping dict
    mapping_dict = {m["template_header"]: m["standard_field"] for m in field_mapping if m.get("standard_field")}
    
    # Fill rows
    filled_rows = []
    for resident in residents:
        row = {}
        for header in template_headers:
            std_field = mapping_dict.get(header, "")
            if std_field:
                row[header] = get_resident_field_value(resident, std_field)
            else:
                row[header] = ""  # Unmapped fields are empty
        filled_rows.append(row)
    
    return filled_rows


def generate_excel(
    headers: List[str],
    rows: List[Dict[str, str]],
    output_path: str,
    sheet_name: str = "Sheet1"
) -> str:
    """Generate Excel file from headers and rows. Returns file path."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Header style
        header_fill = PatternFill(start_color="2B5C8F", end_color="2B5C8F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for i, header in enumerate(headers):
            cell = ws.cell(row=1, column=i + 1, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        # Write data
        for r_idx, row in enumerate(rows):
            for c_idx, header in enumerate(headers):
                cell = ws.cell(row=r_idx + 2, column=c_idx + 1, value=row.get(header, ""))
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
        
        # Auto-adjust column widths
        for i, header in enumerate(headers):
            max_width = len(str(header))
            for row in rows[:50]:  # Sample for width calculation
                val_len = len(str(row.get(header, "")))
                if val_len > max_width:
                    max_width = min(val_len, 30)
            ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = max_width + 4
        
        wb.save(output_path)
        return output_path
    except ImportError:
        raise ImportError("openpyxl required for Excel generation")


def generate_csv(
    headers: List[str],
    rows: List[Dict[str, str]],
    output_path: str
) -> str:
    """Generate CSV file from headers and rows."""
    import csv
    
    with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)
    
    return output_path
