"""
Table Generator - 自动生成社区工作各类Excel表格
读取居民数据库，自动填入真实数据，生成美观的Excel表格
"""
import os
import uuid
import re
from datetime import datetime
from typing import List, Dict, Any
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.models import ResidentMaster, Housing
from app.encryption import decrypt_field


def _get_field(resident, encrypted_field, masked_field):
    """Get real value from encrypted field, fallback to masked."""
    try:
        val = getattr(resident, encrypted_field, None)
        if val:
            decrypted = decrypt_field(val)
            if decrypted:
                return decrypted
    except Exception:
        pass
    # Fallback to masked or empty
    return getattr(resident, masked_field, None) or ""

# 表格保存目录 - 与 assistant.py 的 REPORTS_DIR 保持一致
# assistant.py: os.path.join(os.path.dirname(__file__), "..", "..", "reports") 
#   = backend/app/routers/../.. = backend/reports
# table_generator.py 需要从 backend/app 出发多走一层 ..
TABLES_DIR = os.path.join(os.path.dirname(__file__), "..", "reports")
os.makedirs(TABLES_DIR, exist_ok=True)

# ── 样式定义 ──────────────────────────────────

HEADER_FILL = PatternFill(start_color="2B5C8F", end_color="2B5C8F", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="微软雅黑", size=10)
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = BORDER


def _style_data(cell, align=LEFT_ALIGN):
    cell.font = DATA_FONT
    cell.alignment = align
    cell.border = BORDER


def _auto_width(worksheet):
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


# ── 表格定义 ──────────────────────────────────

TABLE_SCHEMAS = {
    "resident_register": {
        "name": "居民信息登记表",
        "description": "社区所有居民的基本信息登记",
        "filter": None,
        "columns": [
            ("序号", lambda r, i: i + 1, 8),
            ("姓名", lambda r, i: _get_field(r, "name_encrypted", "name_masked"), 12),
            ("性别", lambda r, i: r.gender or "", 8),
            ("年龄", lambda r, i: r.age or "", 8),
            ("民族", lambda r, i: r.ethnicity or "", 10),
            ("身份证号", lambda r, i: _get_field(r, "id_card_encrypted", "id_card_masked"), 20),
            ("联系电话", lambda r, i: _get_field(r, "phone_encrypted", "phone_masked"), 15),
            ("户籍地址", lambda r, i: r.household_address or "", 30),
            ("居住地址", lambda r, i: r.residence_address or "", 30),
            ("网格", lambda r, i: r.grid_name or "", 15),
            ("楼栋单元", lambda r, i: r.building_unit or "", 15),
            ("婚姻状况", lambda r, i: r.marital_status or "", 10),
            ("就业状态", lambda r, i: r.employment_status or "", 12),
            ("医保类型", lambda r, i: r.medical_insurance or "", 12),
            ("登记日期", lambda r, i: r.created_at.strftime("%Y-%m-%d") if r.created_at else "", 12),
            ("备注", lambda r, i: "", 20),
        ],
        "dropdowns": {
            "C": '"男,女"',
            "L": '"未婚,已婚,离异,丧偶"',
        },
    },
    "elderly_visit": {
        "name": "独居老人走访记录表",
        "description": "60岁以上独居老人走访登记",
        "filter": lambda q: q.filter(ResidentMaster.age >= 60, ResidentMaster.is_living_alone == True),
        "columns": [
            ("序号", lambda r, i: i + 1, 8),
            ("姓名", lambda r, i: _get_field(r, "name_encrypted", "name_masked"), 12),
            ("性别", lambda r, i: r.gender or "", 8),
            ("年龄", lambda r, i: r.age or "", 8),
            ("联系电话", lambda r, i: _get_field(r, "phone_encrypted", "phone_masked"), 15),
            ("居住地址", lambda r, i: r.residence_address or "", 30),
            ("网格", lambda r, i: r.grid_name or "", 15),
            ("健康状况", lambda r, i: "", 15),
            ("走访日期", lambda r, i: "", 12),
            ("走访人", lambda r, i: "", 12),
            ("走访内容", lambda r, i: "", 30),
            ("存在问题", lambda r, i: "", 25),
            ("处理措施", lambda r, i: "", 25),
            ("下次走访日期", lambda r, i: "", 14),
            ("备注", lambda r, i: "", 20),
        ],
    },
    "low_income": {
        "name": "低保户调查表",
        "description": "低保户家庭情况调查",
        "filter": lambda q: q.filter(ResidentMaster.is_low_income == True),
        "columns": [
            ("序号", lambda r, i: i + 1, 8),
            ("户主姓名", lambda r, i: r.name_masked or "", 12),
            ("性别", lambda r, i: r.gender or "", 8),
            ("年龄", lambda r, i: r.age or "", 8),
            ("联系电话", lambda r, i: _get_field(r, "phone_encrypted", "phone_masked"), 15),
            ("居住地址", lambda r, i: r.residence_address or "", 30),
            ("网格", lambda r, i: r.grid_name or "", 15),
            ("家庭人口", lambda r, i: "", 10),
            ("低保类型", lambda r, i: "", 12),
            ("享受金额(元/月)", lambda r, i: "", 15),
            ("困难原因", lambda r, i: "", 25),
            ("帮扶措施", lambda r, i: "", 25),
            ("调查日期", lambda r, i: "", 12),
            ("调查人", lambda r, i: "", 12),
            ("备注", lambda r, i: "", 20),
        ],
    },
    "disabled_register": {
        "name": "残疾人登记表",
        "description": "社区残疾人基本信息登记",
        "filter": lambda q: q.filter(ResidentMaster.is_disabled == True),
        "columns": [
            ("序号", lambda r, i: i + 1, 8),
            ("姓名", lambda r, i: _get_field(r, "name_encrypted", "name_masked"), 12),
            ("性别", lambda r, i: r.gender or "", 8),
            ("年龄", lambda r, i: r.age or "", 8),
            ("残疾类别", lambda r, i: r.disability_type or "", 15),
            ("残疾等级", lambda r, i: "", 10),
            ("联系电话", lambda r, i: _get_field(r, "phone_encrypted", "phone_masked"), 15),
            ("居住地址", lambda r, i: r.residence_address or "", 30),
            ("网格", lambda r, i: r.grid_name or "", 15),
            ("是否享受补贴", lambda r, i: "", 12),
            ("监护人姓名", lambda r, i: "", 12),
            ("监护人电话", lambda r, i: "", 15),
            ("登记日期", lambda r, i: "", 12),
            ("备注", lambda r, i: "", 20),
        ],
    },
    "key_population": {
        "name": "重点人员台账",
        "description": "重点人群管理台账",
        "filter": lambda q: q.filter(ResidentMaster.is_key_population == True),
        "columns": [
            ("序号", lambda r, i: i + 1, 8),
            ("姓名", lambda r, i: _get_field(r, "name_encrypted", "name_masked"), 12),
            ("性别", lambda r, i: r.gender or "", 8),
            ("年龄", lambda r, i: r.age or "", 8),
            ("人员类别", lambda r, i: r.key_population_type or "", 15),
            ("联系电话", lambda r, i: _get_field(r, "phone_encrypted", "phone_masked"), 15),
            ("居住地址", lambda r, i: r.residence_address or "", 30),
            ("网格", lambda r, i: r.grid_name or "", 15),
            ("管控等级", lambda r, i: "", 10),
            ("管控措施", lambda r, i: "", 25),
            ("责任人", lambda r, i: "", 12),
            ("走访记录", lambda r, i: "", 30),
            ("登记日期", lambda r, i: "", 12),
            ("备注", lambda r, i: "", 20),
        ],
    },
    "grid_statistics": {
        "name": "网格人员统计表",
        "description": "按网格统计居民分布情况",
        "is_summary": True,  # 特殊类型：汇总统计
        "columns": [],  # 动态生成
    },
    "housing_register": {
        "name": "房屋信息登记表",
        "description": "社区房屋及住户信息登记",
        "is_housing": True,  # 特殊类型：房屋数据
        "columns": [],
    },
    "safety_inspection": {
        "name": "安全巡查记录表",
        "description": "社区安全巡查登记",
        "filter": None,
        "columns": [
            ("序号", lambda r, i: i + 1, 8),
            ("巡查日期", lambda r, i: "", 12),
            ("巡查区域", lambda r, i: r.grid_name or "", 15),
            ("巡查楼栋", lambda r, i: r.building_unit or "", 15),
            ("巡查人", lambda r, i: "", 12),
            ("巡查项目", lambda r, i: "", 20),
            ("发现问题", lambda r, i: "", 30),
            ("整改建议", lambda r, i: "", 25),
            ("整改期限", lambda r, i: "", 12),
            ("整改结果", lambda r, i: "", 20),
            ("复查人", lambda r, i: "", 12),
            ("复查日期", lambda r, i: "", 12),
            ("备注", lambda r, i: "", 20),
        ],
    },
    "appeal_register": {
        "name": "居民诉求登记表",
        "description": "居民诉求及处理记录",
        "filter": None,
        "columns": [
            ("序号", lambda r, i: i + 1, 8),
            ("登记日期", lambda r, i: r.created_at.strftime("%Y-%m-%d") if r.created_at else "", 12),
            ("诉求人姓名", lambda r, i: r.name_masked or "", 12),
            ("联系电话", lambda r, i: _get_field(r, "phone_encrypted", "phone_masked"), 15),
            ("居住地址", lambda r, i: r.residence_address or "", 30),
            ("网格", lambda r, i: r.grid_name or "", 15),
            ("诉求类型", lambda r, i: "", 12),
            ("诉求内容", lambda r, i: "", 30),
            ("受理人", lambda r, i: "", 12),
            ("处理措施", lambda r, i: "", 25),
            ("处理结果", lambda r, i: "", 20),
            ("完成日期", lambda r, i: "", 12),
            ("满意度", lambda r, i: "", 10),
            ("备注", lambda r, i: "", 20),
        ],
    },
    "negotiation_record": {
        "name": "协商会议记录表",
        "description": "协商议事会议记录",
        "filter": None,
        "columns": [
            ("序号", lambda r, i: i + 1, 8),
            ("会议日期", lambda r, i: "", 12),
            ("会议地点", lambda r, i: "", 20),
            ("议题", lambda r, i: "", 30),
            ("主持人", lambda r, i: "", 12),
            ("记录人", lambda r, i: "", 12),
            ("参会人员", lambda r, i: "", 30),
            ("议题说明", lambda r, i: "", 30),
            ("各方意见", lambda r, i: "", 30),
            ("协商结果", lambda r, i: "", 30),
            ("落实措施", lambda r, i: "", 25),
            ("责任人", lambda r, i: "", 12),
            ("完成期限", lambda r, i: "", 12),
            ("备注", lambda r, i: "", 20),
        ],
    },
}


def _get_schema(table_type: str, custom_title: str = "") -> dict:
    """Get schema for table type. Falls back to generic template for unknown types."""
    schema = TABLE_SCHEMAS.get(table_type)
    if schema:
        return schema
    # Unknown type: create a generic template with basic resident info
    title = custom_title if custom_title else f"{table_type}登记表"
    return {
        "name": title,
        "description": f"自定义表格（{table_type}）",
        "filter": None,
        "columns": [
            ("序号", lambda r, i: i + 1, 8),
            ("姓名", lambda r, i: _get_field(r, "name_encrypted", "name_masked"), 12),
            ("性别", lambda r, i: r.gender or "", 8),
            ("年龄", lambda r, i: r.age or "", 8),
            ("联系电话", lambda r, i: _get_field(r, "phone_encrypted", "phone_masked"), 15),
            ("居住地址", lambda r, i: r.residence_address or "", 30),
            ("网格", lambda r, i: r.grid_name or "", 15),
            ("楼栋单元", lambda r, i: r.building_unit or "", 15),
            ("备注1", lambda r, i: "", 15),
            ("备注2", lambda r, i: "", 15),
            ("备注3", lambda r, i: "", 15),
            ("备注4", lambda r, i: "", 15),
            ("备注5", lambda r, i: "", 15),
            ("备注", lambda r, i: "", 20),
        ],
    }


def generate_table(table_type: str, db: Session, custom_title: str = "") -> str:
    """Generate an Excel table of the given type. Returns file path."""
    schema = _get_schema(table_type, custom_title)

    wb = Workbook()
    ws = wb.active
    ws.title = schema["name"]

    # 根据实际列数计算合并范围
    num_cols = len(schema.get("columns", [])) if not schema.get("is_summary") and not schema.get("is_housing") else 12
    last_col_letter = get_column_letter(num_cols)

    # 标题行
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws["A1"]
    title_cell.value = schema["name"]
    title_cell.font = Font(name="微软雅黑", size=16, bold=True, color="2B5C8F")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # 副标题
    ws.merge_cells(f"A2:{last_col_letter}2")
    subtitle_cell = ws["A2"]
    subtitle_cell.value = f"生成日期：{datetime.now().strftime('%Y年%m月%d日')}    数据来源：数邻智算系统"
    subtitle_cell.font = Font(name="微软雅黑", size=9, color="888888")
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # 处理特殊类型
    if schema.get("is_summary"):
        _generate_grid_summary(ws, db)
    elif schema.get("is_housing"):
        _generate_housing_table(ws, db)
    else:
        _generate_data_table(ws, db, schema)

    # 保存
    filename = f"table_{table_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:6]}.xlsx"
    file_path = os.path.join(TABLES_DIR, filename)
    wb.save(file_path)
    return file_path


def _generate_data_table(ws, db: Session, schema: dict):
    """Generate a standard data table with resident records."""
    # Query all residents - fill in what we have, leave blank what we don't
    residents = db.query(ResidentMaster).order_by(ResidentMaster.id).limit(5000).all()

    # Header row
    header_row = 3
    ws.row_dimensions[header_row].height = 28
    for col_idx, (col_name, _, _) in enumerate(schema["columns"], 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        _style_header(cell)

    # Freeze panes
    ws.freeze_panes = f"A{header_row + 1}"

    # Data rows
    for i, resident in enumerate(residents):
        row_num = header_row + 1 + i
        ws.row_dimensions[row_num].height = 22
        for col_idx, (_, getter, _) in enumerate(schema["columns"], 1):
            value = getter(resident, i)
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            _style_data(cell, CENTER_ALIGN if col_idx <= 2 else LEFT_ALIGN)

        # 交替行颜色
        if i % 2 == 1:
            for col_idx in range(1, len(schema["columns"]) + 1):
                ws.cell(row=row_num, column=col_idx).fill = PatternFill(
                    start_color="F5F8FA", end_color="F5F8FA", fill_type="solid"
                )

    # 下拉选项
    for col_letter, formula in schema.get("dropdowns", {}).items():
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "请从下拉列表中选择"
        dv.errorTitle = "输入错误"
        dv.prompt = "点击选择"
        dv.promptTitle = "选项"
        last_row = header_row + len(residents) + 100
        dv.add(f"{col_letter}{header_row + 1}:{col_letter}{last_row}")
        ws.add_data_validation(dv)

    # 自动列宽
    for col_idx, (_, _, width) in enumerate(schema["columns"], 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 如果没有数据，添加示例行
    if not residents:
        row_num = header_row + 1
        for col_idx, (_, _, _) in enumerate(schema["columns"], 1):
            cell = ws.cell(row=row_num, column=col_idx, value="" if col_idx > 1 else "（暂无数据，请上传居民信息后重新生成）")
            _style_data(cell)
        ws.merge_cells(f"A{row_num}:{get_column_letter(len(schema['columns']))}{row_num}")


def _generate_grid_summary(ws, db: Session):
    """Generate grid statistics summary table."""
    from sqlalchemy import func

    # Header
    header_row = 3
    headers = ["序号", "网格名称", "居民总数", "男性", "女性", "60岁以上", "独居老人",
               "残疾人", "低保户", "留守儿童", "重点人群", "备注"]
    ws.row_dimensions[header_row].height = 28
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        _style_header(cell)

    ws.freeze_panes = f"A{header_row + 1}"

    # Query grid stats
    grids = db.query(ResidentMaster.grid_name).distinct().filter(
        ResidentMaster.grid_name.isnot(None)
    ).all()

    for i, (grid_name,) in enumerate(sorted(g[0] for g in grids if g[0])):
        row_num = header_row + 1 + i
        ws.row_dimensions[row_num].height = 22

        total = db.query(ResidentMaster).filter(ResidentMaster.grid_name == grid_name).count()
        male = db.query(ResidentMaster).filter(ResidentMaster.grid_name == grid_name, ResidentMaster.gender == "男").count()
        female = db.query(ResidentMaster).filter(ResidentMaster.grid_name == grid_name, ResidentMaster.gender == "女").count()
        elderly = db.query(ResidentMaster).filter(ResidentMaster.grid_name == grid_name, ResidentMaster.age >= 60).count()
        alone = db.query(ResidentMaster).filter(ResidentMaster.grid_name == grid_name, ResidentMaster.is_living_alone == True, ResidentMaster.age >= 60).count()
        disabled = db.query(ResidentMaster).filter(ResidentMaster.grid_name == grid_name, ResidentMaster.is_disabled == True).count()
        low_income = db.query(ResidentMaster).filter(ResidentMaster.grid_name == grid_name, ResidentMaster.is_low_income == True).count()
        left_behind = db.query(ResidentMaster).filter(ResidentMaster.grid_name == grid_name, ResidentMaster.is_left_behind_child == True).count()
        key_pop = db.query(ResidentMaster).filter(ResidentMaster.grid_name == grid_name, ResidentMaster.is_key_population == True).count()

        values = [i + 1, grid_name, total, male, female, elderly, alone, disabled, low_income, left_behind, key_pop, ""]
        for col_idx, v in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=v)
            _style_data(cell, CENTER_ALIGN)

        if i % 2 == 1:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_idx).fill = PatternFill(
                    start_color="F5F8FA", end_color="F5F8FA", fill_type="solid"
                )

    # 合计行
    total_row = header_row + 1 + len([g for g in grids if g[0]])
    ws.cell(row=total_row, column=1, value="").border = BORDER
    ws.cell(row=total_row, column=2, value="合计").font = Font(name="微软雅黑", size=10, bold=True)
    ws.cell(row=total_row, column=2).alignment = CENTER_ALIGN
    ws.cell(row=total_row, column=2).border = BORDER
    for col_idx in range(3, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        formula = f"=SUM({col_letter}{header_row + 1}:{col_letter}{total_row - 1})"
        cell = ws.cell(row=total_row, column=col_idx, value=formula)
        cell.font = Font(name="微软雅黑", size=10, bold=True)
        cell.alignment = CENTER_ALIGN
        cell.border = BORDER
        cell.fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

    for col_idx, w in enumerate([8, 20, 12, 10, 10, 12, 12, 10, 10, 12, 12, 20], 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def _generate_housing_table(ws, db: Session):
    """Generate housing registration table."""
    header_row = 3
    headers = ["序号", "房屋地址", "楼栋", "单元", "房号", "房屋类型",
               "面积(m²)", "业主姓名", "业主电话", "居住人数", "网格", "备注"]
    ws.row_dimensions[header_row].height = 28
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        _style_header(cell)

    ws.freeze_panes = f"A{header_row + 1}"

    houses = db.query(Housing).order_by(Housing.id).limit(5000).all()
    for i, h in enumerate(houses):
        row_num = header_row + 1 + i
        ws.row_dimensions[row_num].height = 22
        values = [
            i + 1, h.address, h.building_name, h.unit_number, h.room_number,
            h.housing_type, h.area_sqm, h.owner_name, h.owner_phone,
            h.resident_count, h.grid_name, ""
        ]
        for col_idx, v in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=v)
            _style_data(cell, CENTER_ALIGN if col_idx <= 2 else LEFT_ALIGN)

        if i % 2 == 1:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_idx).fill = PatternFill(
                    start_color="F5F8FA", end_color="F5F8FA", fill_type="solid"
                )

    for col_idx, w in enumerate([8, 30, 12, 10, 10, 12, 12, 12, 15, 10, 15, 20], 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def get_table_types():
    return [{"type": k, "name": s["name"], "description": s["description"]} for k, s in TABLE_SCHEMAS.items()]
